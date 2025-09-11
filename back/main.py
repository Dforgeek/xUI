import enum
from datetime import datetime, timezone
from typing import List, Optional, Literal, Dict, Any, Set

from fastapi import FastAPI, UploadFile, File, Depends, HTTPException, Query
from fastapi import status
from fastapi.middleware.cors import CORSMiddleware
from fastapi.security import APIKeyHeader

from pydantic import BaseModel, Field, field_validator
from pydantic_settings import BaseSettings
import os

from sqlalchemy import (
    BigInteger, SmallInteger, String, Text, ARRAY, Integer, ForeignKey,
    Index, UniqueConstraint, select, func, and_, literal_column, Boolean, DateTime
)
from sqlalchemy.ext.asyncio import create_async_engine, async_sessionmaker, AsyncSession
from sqlalchemy.orm import DeclarativeBase, mapped_column, Mapped, relationship
from sqlalchemy import update

# =========================
# Settings
# =========================

class Settings(BaseSettings):
    DATABASE_URL: str = os.getenv("DATABASE_URL")
    CORS_ALLOW_ORIGINS: List[str] = ["*"]

settings = Settings()

# =========================
# DB / ORM
# =========================

class Base(DeclarativeBase):
    pass

class UserInfo(Base):
    __tablename__ = "user_info"
    id: Mapped[int] = mapped_column(BigInteger, primary_key=True, autoincrement=True)
    telegram_id: Mapped[int] = mapped_column(BigInteger, nullable=False, index=True)
    post: Mapped[int] = mapped_column(Integer, nullable=False)
    command_id: Mapped[int] = mapped_column(BigInteger, nullable=False, index=True)

    # NEW personal fields (nullable, can be backfilled later)
    first_name: Mapped[Optional[str]] = mapped_column(String(100), nullable=True)
    last_name:  Mapped[Optional[str]] = mapped_column(String(100), nullable=True)
    email:      Mapped[Optional[str]] = mapped_column(String(255), nullable=True)
    telegram:   Mapped[Optional[str]] = mapped_column(String(64),  nullable=True)

    surveys_subject: Mapped[list["Survey"]] = relationship(
        back_populates="subject_user",
        foreign_keys=lambda: [Survey.subject_user_id],
        cascade="all,delete",
        passive_deletes=True,
    )
    survey_responses: Mapped[list["SurveyRespondent"]] = relationship(
        back_populates="user",
        cascade="all, delete-orphan",
        passive_deletes=True,
    )
    answers: Mapped[list["SurveyAnswer"]] = relationship(
        back_populates="user",
        cascade="all, delete-orphan",
        passive_deletes=True,
    )
class Block(Base):
    __tablename__ = "block"
    id: Mapped[int] = mapped_column(BigInteger, primary_key=True, autoincrement=True)
    block_name: Mapped[str] = mapped_column(String(255), nullable=False, unique=True)

    questions: Mapped[list["Question"]] = relationship(
        back_populates="block", cascade="all, delete-orphan", passive_deletes=True
    )


class Question(Base):
    __tablename__ = "question"
    id: Mapped[int] = mapped_column(BigInteger, primary_key=True, autoincrement=True)
    block_id: Mapped[int] = mapped_column(
        BigInteger,
        ForeignKey("block.id", ondelete="CASCADE"),
        nullable=False,
        index=True,
    )
    question_text: Mapped[str] = mapped_column(Text, nullable=False)
    question_type: Mapped[int] = mapped_column(SmallInteger, nullable=False)
    answer_fields: Mapped[str] = mapped_column(Text, nullable=False)

    block: Mapped[Block] = relationship(back_populates="questions")
    in_surveys: Mapped[list["SurveyQuestion"]] = relationship(
        back_populates="question", cascade="all,delete", passive_deletes=True
    )
    answers: Mapped[list["SurveyAnswer"]] = relationship(
        back_populates="question", cascade="all,delete", passive_deletes=True
    )

class SurveyPreset(Base):
    __tablename__ = "survey_preset"
    id: Mapped[int] = mapped_column(BigInteger, primary_key=True, autoincrement=True)
    questions: Mapped[list[int]] = mapped_column(ARRAY(BigInteger), nullable=False)

class Survey(Base):
    __tablename__ = "survey"
    id: Mapped[int] = mapped_column(BigInteger, primary_key=True, autoincrement=True)

    subject_user_id: Mapped[int] = mapped_column(
        BigInteger, ForeignKey("user_info.id", ondelete="RESTRICT"),
        nullable=False, index=True
    )
    respondent_user_id: Mapped[int] = mapped_column(
        BigInteger, ForeignKey("user_info.id", ondelete="RESTRICT"),
        nullable=False, index=True
    )

    created_at: Mapped[datetime] = mapped_column(DateTime(timezone=True), nullable=False)
    deadline:   Mapped[datetime] = mapped_column(DateTime(timezone=True), nullable=False)
    notifications_before: Mapped[int] = mapped_column(BigInteger, nullable=False)
    anonymous: Mapped[bool] = mapped_column(Boolean, nullable=False, default=False)
    review_type: Mapped[str] = mapped_column(String(10), nullable=False)
    title: Mapped[Optional[str]] = mapped_column(String(255), nullable=True)  # already added earlier

    subject_user: Mapped["UserInfo"] = relationship(
        back_populates="surveys_subject", foreign_keys=[subject_user_id]
    )
    respondent_user: Mapped["UserInfo"] = relationship(
        foreign_keys=[respondent_user_id], lazy="joined"
    )

    respondents: Mapped[list["SurveyRespondent"]] = relationship(
        back_populates="survey", cascade="all, delete-orphan", passive_deletes=True
    )
    answers: Mapped[list["SurveyAnswer"]] = relationship(
        back_populates="survey", cascade="all, delete-orphan", passive_deletes=True
    )
    questions: Mapped[list["SurveyQuestion"]] = relationship(
        back_populates="survey", cascade="all, delete-orphan", passive_deletes=True
    )


class SurveyQuestion(Base):
    __tablename__ = "survey_question"
    id: Mapped[int] = mapped_column(BigInteger, primary_key=True, autoincrement=True)
    question_id: Mapped[int] = mapped_column(
        BigInteger, ForeignKey("question.id", ondelete="CASCADE"), nullable=False
    )
    survey_id: Mapped[int] = mapped_column(
        BigInteger, ForeignKey("survey.id", ondelete="CASCADE"), nullable=False
    )
    __table_args__ = (UniqueConstraint("question_id", "survey_id", name="survey_question_survey_id_question_id_idx"),)

    # NEW: optional flag
    optional: Mapped[bool] = mapped_column(Boolean, nullable=False, default=False)

    survey: Mapped["Survey"] = relationship(back_populates="questions")
    question: Mapped["Question"] = relationship(back_populates="in_surveys")

class SurveyRespondent(Base):
    __tablename__ = "survey_respondent"
    id: Mapped[int] = mapped_column(BigInteger, primary_key=True, autoincrement=True)
    user_id: Mapped[int] = mapped_column(
        BigInteger, ForeignKey("user_info.id", ondelete="CASCADE"), nullable=False
    )
    survey_id: Mapped[int] = mapped_column(
        BigInteger, ForeignKey("survey.id", ondelete="CASCADE"), nullable=False
    )
    __table_args__ = (
        UniqueConstraint("user_id", "survey_id", name="survey_respondent_user_id_survey_id_idx"),
    )

    survey: Mapped["Survey"] = relationship(back_populates="respondents")
    # This expects UserInfo.survey_responses to exist:
    user: Mapped["UserInfo"] = relationship(back_populates="survey_responses")

class SurveyAnswer(Base):
    __tablename__ = "survey_answer"
    id: Mapped[int] = mapped_column(BigInteger, primary_key=True, autoincrement=True)
    survey_id: Mapped[int] = mapped_column(
        BigInteger, ForeignKey("survey.id", ondelete="CASCADE"), nullable=False, index=True
    )
    user_id: Mapped[int] = mapped_column(
        BigInteger, ForeignKey("user_info.id", ondelete="CASCADE"), nullable=False
    )
    question_id: Mapped[int] = mapped_column(
        BigInteger, ForeignKey("question.id", ondelete="CASCADE"), nullable=False, index=True
    )
    answer: Mapped[str] = mapped_column(Text, nullable=False)

    __table_args__ = (
        UniqueConstraint("survey_id", "user_id", "question_id", name="uq_answer_triplet"),
        Index("ix_answer_survey_q", "survey_id", "question_id"),
    )

    survey: Mapped["Survey"] = relationship(back_populates="answers")
    user:   Mapped["UserInfo"] = relationship(back_populates="answers")
    question: Mapped["Question"] = relationship(back_populates="answers")

# =========================
# Engine / Session
# =========================

engine = create_async_engine(settings.DATABASE_URL, echo=False, pool_pre_ping=True)
SessionLocal = async_sessionmaker(engine, expire_on_commit=False)

async def get_session() -> AsyncSession:
    async with SessionLocal() as session:
        yield session

# =========================
# Schemas (Pydantic)
# =========================

class EmployeeIn(BaseModel):
    telegram_id: int
    post: int
    command_id: int
    first_name: Optional[str] = None
    last_name: Optional[str] = None
    email: Optional[str] = None
    telegram: Optional[str] = None

class Employee(EmployeeIn):
    id: int

class BlockIn(BaseModel):
    block_name: str

class BlockOut(BlockIn):
    id: int

class QuestionIn(BaseModel):
    block_id: int
    question_text: str
    question_type: int = Field(..., description="Например: 0=текст, 1=шкала, 2=множественный выбор")
    answer_fields: str = Field(..., description="Схема полей/вариантов в JSON/CSV и т.п.")

class QuestionOut(QuestionIn):
    id: int

class PresetIn(BaseModel):
    questions: List[int]

class PresetOut(PresetIn):
    id: int

ReviewType = Literal["180", "360"]

class InitiateSurveyIn(BaseModel):
    subject_user_id: int
    reviewer_user_ids: List[int] = Field(default_factory=list)
    review_type: ReviewType
    question_ids: Optional[List[int]] = None
    deadline: datetime
    notifications_before: int = 0
    anonymous: bool = False
    title: Optional[str] = None  # optional survey title (shown in envelope)

    @field_validator("deadline")
    @classmethod
    def ensure_deadline_tz(cls, v: datetime) -> datetime:
        # ensure aware (UTC) to match timestamptz
        if v.tzinfo is None or v.tzinfo.utcoffset(v) is None:
            return v.replace(tzinfo=timezone.utc)
        return v
    
class InitiatedPersonalSurvey(BaseModel):
    surveyId: str
    respondent_user_id: int
    linkToken: str

class InitiateSurveyBatchOut(BaseModel):
    batch_created: List[InitiatedPersonalSurvey]
    questions_count: int
    
class SurveyOut(BaseModel):
    id: int
    subject_user_id: int
    created_at: datetime
    deadline: datetime
    notifications_before: int
    anonymous: bool
    review_type: ReviewType
    participants_count: int
    questions_count: int

class SurveyListItem(BaseModel):
    id: int
    created_at: datetime
    deadline: datetime
    participants_count: int
    completed_count: int
    status: Literal["pending", "in_progress", "completed"]
    reviewer_user_ids: List[int]

class AnswerIn(BaseModel):
    user_id: int
    question_id: int
    answer: str

class QuestionAggregate(BaseModel):
    question_id: int
    question_text: str
    answers: List[Dict[str, Any]]

class SurveyAggregateOut(BaseModel):
    survey_id: int
    anonymous: bool
    review_type: ReviewType
    total_respondents: int
    responded: int
    by_question: List[QuestionAggregate]

class BulkAnswerItem(BaseModel):
    question_id: int
    answer: str

class BulkAnswersIn(BaseModel):
    user_id: int
    answers: List[BulkAnswerItem]

class SurveyInfoOut(BaseModel):
    subject_user_id: int
    created_at: datetime
    deadline: datetime
    anonymous: bool
    review_type: ReviewType

class QuestionFormItem(BaseModel):
    question_id: int
    question_text: str
    question_type: int
    answer_fields: str

class SurveyFormOut(BaseModel):
    user_id: int
    survey_id: int
    survey_info: SurveyInfoOut
    questions: List[QuestionFormItem]

# =========================
# App
# =========================

app = FastAPI(title="360 Survey Backend", version="0.1.0")
try:
    from frontend_api import v1 as frontend_v1
    app.include_router(frontend_v1)  # ensures router mounts on import
except Exception as e:
    print("Failed to load one_block_api:", e)
app.add_middleware(
    CORSMiddleware,
    allow_origins=settings.CORS_ALLOW_ORIGINS,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# =========================
# Employees
# =========================

@app.post("/employees", response_model=Employee, status_code=201)
async def create_employee(payload: EmployeeIn, db: AsyncSession = Depends(get_session)):
    obj = UserInfo(**payload.model_dump())
    db.add(obj)
    await db.commit()
    await db.refresh(obj)
    return Employee(id=obj.id, **payload.model_dump())

@app.get("/employees", response_model=List[Employee])
async def list_employees(db: AsyncSession = Depends(get_session), limit: int = Query(100, ge=1, le=1000), offset: int = Query(0, ge=0)):
    res = await db.execute(select(UserInfo).limit(limit).offset(offset))
    rows = res.scalars().all()
    return [
        Employee(
            id=r.id, telegram_id=r.telegram_id, post=r.post, command_id=r.command_id,
            first_name=r.first_name, last_name=r.last_name, email=r.email, telegram=r.telegram
        ) for r in rows
    ]

@app.put("/employees/{employee_id}", response_model=Employee)
async def update_employee(employee_id: int, payload: EmployeeIn, db: AsyncSession = Depends(get_session)):
    res = await db.execute(select(UserInfo).where(UserInfo.id == employee_id))
    obj = res.scalar_one_or_none()
    if not obj:
        raise HTTPException(404, "Employee not found")
    for k, v in payload.model_dump().items():
        setattr(obj, k, v)
    await db.commit()
    await db.refresh(obj)
    return Employee(id=obj.id, **payload.model_dump())

@app.delete("/employees/{employee_id}", status_code=204)
async def delete_employee(employee_id: int, db: AsyncSession = Depends(get_session)):
    res = await db.execute(select(UserInfo).where(UserInfo.id == employee_id))
    obj = res.scalar_one_or_none()
    if not obj:
        raise HTTPException(404, "Employee not found")
    await db.delete(obj)
    await db.commit()
    return

# = Excel import =
from openpyxl import load_workbook
@app.post("/employees/import-xlsx", response_model=Dict[str, int])
async def import_employees_xlsx(file: UploadFile = File(...), db: AsyncSession = Depends(get_session)):
    if not file.filename.lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(400, "Upload an .xlsx file")
    wb = load_workbook(filename=file.file, read_only=True)
    ws = wb.active
    # ожидаем столбцы: telegram_id, post, command_id (в первой строке — заголовки)
    headers = [str(c.value).strip() if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
    try:
        idx_tg = headers.index("telegram_id")
        idx_post = headers.index("post")
        idx_cmd = headers.index("command_id")
    except ValueError:
        raise HTTPException(400, "Header row must contain telegram_id, post, command_id")
    created = 0
    for row in ws.iter_rows(min_row=2):
        tg = row[idx_tg].value
        post = row[idx_post].value
        cmd = row[idx_cmd].value
        if tg is None or post is None or cmd is None:
            continue
        db.add(UserInfo(telegram_id=int(tg), post=int(post), command_id=int(cmd)))
        created += 1
    await db.commit()
    return {"created": created}

# =========================
# Blocks & Questions & Presets (управление библиотекой)
# =========================

@app.post("/blocks", response_model=BlockOut, status_code=201)
async def create_block(payload: BlockIn, db: AsyncSession = Depends(get_session)):
    obj = Block(**payload.model_dump())
    db.add(obj)
    await db.commit()
    await db.refresh(obj)
    return BlockOut(id=obj.id, **payload.model_dump())

@app.get("/blocks", response_model=List[BlockOut])
async def list_blocks(db: AsyncSession = Depends(get_session)):
    res = await db.execute(select(Block))
    return [BlockOut(id=b.id, block_name=b.block_name) for b in res.scalars().all()]

@app.post("/questions", response_model=QuestionOut, status_code=201)
async def create_question(payload: QuestionIn, db: AsyncSession = Depends(get_session)):
    # ensure block exists
    exists = await db.scalar(select(func.count()).select_from(Block).where(Block.id == payload.block_id))
    if not exists:
        raise HTTPException(400, "block_id not found")
    obj = Question(**payload.model_dump())
    db.add(obj)
    await db.commit()
    await db.refresh(obj)
    return QuestionOut(id=obj.id, **payload.model_dump())

@app.put("/questions/{question_id}", response_model=QuestionOut)
async def update_question(question_id: int, payload: QuestionIn, db: AsyncSession = Depends(get_session)):
    res = await db.execute(select(Question).where(Question.id == question_id))
    obj = res.scalar_one_or_none()
    if not obj:
        raise HTTPException(404, "Question not found")
    for k, v in payload.model_dump().items():
        setattr(obj, k, v)
    await db.commit()
    await db.refresh(obj)
    return QuestionOut(id=obj.id, **payload.model_dump())

@app.get("/questions", response_model=List[QuestionOut])
async def list_questions(
    db: AsyncSession = Depends(get_session),
    block_id: Optional[int] = None
):
    stmt = select(Question)
    if block_id:
        stmt = stmt.where(Question.block_id == block_id)
    res = await db.execute(stmt)
    return [QuestionOut(id=q.id, block_id=q.block_id, question_text=q.question_text,
                        question_type=q.question_type, answer_fields=q.answer_fields)
            for q in res.scalars().all()]

@app.post("/presets", response_model=PresetOut, status_code=201)
async def create_preset(payload: PresetIn, db: AsyncSession = Depends(get_session)):
    # validate questions exist
    if payload.questions:
        count = await db.scalar(select(func.count()).select_from(Question).where(Question.id.in_(payload.questions)))
        if count != len(payload.questions):
            raise HTTPException(400, "Some question IDs do not exist")
    obj = SurveyPreset(questions=payload.questions)
    db.add(obj)
    await db.commit()
    await db.refresh(obj)
    return PresetOut(id=obj.id, questions=obj.questions)

@app.put("/presets/{preset_id}", response_model=PresetOut)
async def update_preset(preset_id: int, payload: PresetIn, db: AsyncSession = Depends(get_session)):
    res = await db.execute(select(SurveyPreset).where(SurveyPreset.id == preset_id))
    obj = res.scalar_one_or_none()
    if not obj:
        raise HTTPException(404, "Preset not found")
    obj.questions = payload.questions
    await db.commit()
    await db.refresh(obj)
    return PresetOut(id=obj.id, questions=obj.questions)

@app.get("/presets", response_model=List[PresetOut])
async def list_presets(db: AsyncSession = Depends(get_session)):
    res = await db.execute(select(SurveyPreset))
    return [PresetOut(id=p.id, questions=p.questions) for p in res.scalars().all()]

# =========================
# Survey init / list / answers / aggregate
# =========================

async def build_question_ids(
    db: AsyncSession,
    preset_id: Optional[int],
    selected_block_ids: Optional[List[int]],
    additional_question_ids: Optional[List[int]],
) -> List[int]:
    qids: List[int] = []
    if preset_id:
        preset = await db.scalar(select(SurveyPreset).where(SurveyPreset.id == preset_id))
        if not preset:
            raise HTTPException(400, "preset_id not found")
        qids.extend(preset.questions)
    if selected_block_ids:
        res = await db.execute(select(Question.id).where(Question.block_id.in_(selected_block_ids)))
        qids.extend([r[0] for r in res.all()])
    if additional_question_ids:
        qids.extend(additional_question_ids)
    # уникализируем, сохраняем порядок
    seen = set()
    ordered = []
    for q in qids:
        if q not in seen:
            seen.add(q)
            ordered.append(q)
    if not ordered:
        raise HTTPException(400, "No questions selected")
    return ordered





@app.post("/surveys/{survey_id}/answers/bulk", status_code=201)
async def submit_answers_bulk(
    survey_id: int,
    payload: BulkAnswersIn,
    db: AsyncSession = Depends(get_session)
):
    # базовая валидация входных данных
    if not payload.answers:
        raise HTTPException(400, "answers must be a non-empty list")

    # проверка, что пользователь — респондент данного опроса
    belongs = await db.scalar(
        select(func.count())
        .select_from(SurveyRespondent)
        .where(
            and_(
                SurveyRespondent.survey_id == survey_id,
                SurveyRespondent.user_id == payload.user_id
            )
        )
    )
    if not belongs:
        raise HTTPException(400, "User is not a respondent of this survey")

    # собрать список вопросів из payload
    qids = [a.question_id for a in payload.answers]

    # проверить, что ВСЕ эти вопросы входят в опрос
    count_q = await db.scalar(
        select(func.count())
        .select_from(SurveyQuestion)
        .where(
            and_(
                SurveyQuestion.survey_id == survey_id,
                SurveyQuestion.question_id.in_(qids)
            )
        )
    )
    if count_q != len(qids):
        raise HTTPException(400, "Some question_id do not belong to this survey")

    # вытащить уже существующие ответы пользователя по этим вопросам
    existing_res = await db.execute(
        select(SurveyAnswer)
        .where(
            and_(
                SurveyAnswer.survey_id == survey_id,
                SurveyAnswer.user_id == payload.user_id,
                SurveyAnswer.question_id.in_(qids)
            )
        )
    )
    existing = { (row.question_id): row for row in existing_res.scalars().all() }

    created_cnt = 0
    updated_cnt = 0

    # апсерт: обновить существующие, вставить новые
    for item in payload.answers:
        if item.question_id in existing:
            # обновляем текст ответа
            obj = existing[item.question_id]
            obj.answer = item.answer
            updated_cnt += 1
        else:
            db.add(SurveyAnswer(
                survey_id=survey_id,
                user_id=payload.user_id,
                question_id=item.question_id,
                answer=item.answer
            ))
            created_cnt += 1

    await db.commit()
    return {"ok": True, "created": created_cnt, "updated": updated_cnt}

@app.get("/surveys/{survey_id}/form", response_model=SurveyFormOut)
async def get_survey_form(
    survey_id: int,
    user_id: int = Query(..., description="ID пользователя-респондента"),
    db: AsyncSession = Depends(get_session)
):
    # основной объект опроса
    survey = await db.scalar(select(Survey).where(Survey.id == survey_id))
    if not survey:
        raise HTTPException(404, "Survey not found")

    # проверка респондента
    belongs = await db.scalar(
        select(func.count())
        .select_from(SurveyRespondent)
        .where(
            and_(
                SurveyRespondent.survey_id == survey_id,
                SurveyRespondent.user_id == user_id
            )
        )
    )
    if not belongs:
        raise HTTPException(403, "User is not a respondent of this survey")

    # вопросы
    q_res = await db.execute(
        select(
            Question.id,
            Question.question_text,
            Question.question_type,
            Question.answer_fields
        )
        .join(SurveyQuestion, SurveyQuestion.question_id == Question.id)
        .where(SurveyQuestion.survey_id == survey_id)
        .order_by(Question.id)
    )
    questions = [
        QuestionFormItem(
            question_id=qid,
            question_text=qtext,
            question_type=qtype,
            answer_fields=afields
        )
        for (qid, qtext, qtype, afields) in q_res.all()
    ]

    survey_info = SurveyInfoOut(
        subject_user_id=survey.subject_user_id,
        created_at=survey.created_at,
        deadline=survey.deadline,
        anonymous=survey.anonymous,
        review_type=survey.review_type,   # type: ignore
    )

    return SurveyFormOut(
        user_id=user_id,
        survey_id=survey_id,
        survey_info=survey_info,
        questions=questions
    )



# =========================
# Health
# =========================
@app.get("/healthz")
async def health():
    from sqlalchemy.dialects import postgresql
    print("created_at tz? ->", Survey.__table__.c.created_at.type.timezone)  # should print True
    print("deadline   tz? ->", Survey.__table__.c.deadline.type.timezone)    # should print True

    # Optional: see compiled DML cast
    from sqlalchemy import insert
    print(insert(Survey).compile(dialect=postgresql.dialect()))
    return {"status": "ok"}
